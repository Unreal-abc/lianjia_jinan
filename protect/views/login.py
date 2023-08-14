from flask import Blueprint
from flask import render_template
from flask import request,jsonify
from protect import utils
import openpyxl,os
log=Blueprint('log',__name__)
class data():
    name=None
    value=None
@log.route('/',methods=['GET','POST'])
def index():
    return  render_template('bigdata.html')
@log.route('/login',methods=['GET','POST'])
def login():
    if request.method == 'POST':
        user=request.form.get('user')
        password = request.form.get('pwd')
        result=utils.login(user,password)
        name=result['user']
        value = result['value']
        data.name=name
        data.value=value
        if name==str(user) and str(password)==value:
            form={'user':user,'password':password}
    return render_template('login.html', form=form)
@log.route('/res1',methods=['GET','POST'])
def res1():
    return  render_template('res.html')
@log.route('/res',methods=['GET','POST'])
def res():
    if request.method == 'POST':
        user = request.form.get('user')
        password = request.form.get('pwd')
        path = os.getcwd()
        wb=openpyxl.load_workbook(path+'\\protect\\static\\shuju\用户表.xlsx')
        ws=wb.worksheets[0]
        a=ws.max_row
        ws.cell(a+1,1).value=user
        ws.cell(a+1,2).value=password
        wb.save(path+'\\protect\\static\\shuju\用户表.xlsx')
    return  render_template('bigdata.html')
@log.route('/console',methods=['GET','POST'])
def console():
    form={'user':data.name,'password':data.value}
    return render_template('济南地图.html')
@log.route('/setting',methods=['GET','POST'])
def setting():
    return render_template('二手房各区单价排名.html')
@log.route('/pie',methods=['GET','POST'])
def pie():
    return render_template('各区二手房数量占比图.html')
@log.route('/ciyun',methods=['GET','POST'])
def ciyun():
    return render_template('小区词云图.html')
@log.route('/line1',methods=['GET','POST'])
def line1():
    return render_template('年份折线图1.html')
@log.route('/chaoxiang',methods=['GET','POST'])
def chaoxiang():
    return render_template('二手房面积价格散点图.html')
@log.route('/guanzhu',methods=['GET','POST'])
def guanzhu():
    return render_template('二手房各区关注人数排名.html')
@log.route('/yuce',methods=['GET','POST'])
def yuce():
    return render_template('二手房朝向占比图.html')