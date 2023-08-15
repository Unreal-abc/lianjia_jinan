import pymysql
import pyecharts
import openpyxl
from pyecharts import *
from pyecharts import *
import numpy as np
import os
import pandas as pd
from pyecharts import Map, Bar, WordCloud, Scatter, Pie

dict={'lixia':'历下区','laiwuqu':'莱芜区','shizhong':'市中区','tianqiao':'天桥区','licheng':'历城区','huaiyin':"槐荫区",'gaoxin':'高新区','jiyang':'济阳区','shanghe':"商河县",'pingyin':'平阴县','zhangqiu1':'章丘区',
      'changqing':'长清区'}
def login(user,password):
    path = os.getcwd()
    wb = openpyxl.load_workbook(path + '\\protect\\static\\shuju\用户表.xlsx')
    ws = wb.worksheets[0]
    a = ws.max_row
    for i in range(2,a+1):
        if ws.cell(i,1).value==user and ws.cell(i,2).value==password:
            result = {'user': str(ws.cell(i, 1).value),'value':str(ws.cell(i,2).value)}
            return result
def ex():
    path = os.getcwd()
    wb=openpyxl.load_workbook(path+'\\protect\\static\\shuju\hrb.xlsx')
    ws=wb['Sheet1']
    a=2
    db = pymysql.connect(host='192.168.85.100', user='root', password='123456', port=3306, database='lianjia')
    cursor = db.cursor()
    sql = "SELECT *  FROM guangzhou"
    cursor.execute(sql)
    db.close()
    info = cursor.fetchall()
    for i in info:
        ws.cell(a, 1).value=i[0]
        ws.cell(a, 2).value = i[1]
        ws.cell(a, 3).value = i[2]
        ws.cell(a, 4).value = i[3]
        ws.cell(a, 5).value = i[4]
        ws.cell(a, 6).value = i[5]
        ws.cell(a, 7).value = i[6]
        ws.cell(a, 8).value = i[7]
        ws.cell(a, 9).value = i[8]

        ws.cell(a, 10).value = i[9]
        ws.cell(a, 11).value = i[10]
        ws.cell(a, 12).value = i[11]
        ws.cell(a, 13).value = i[12]
        ws.cell(a, 14).value = i[13]
        ws.cell(a, 15).value = dict[i[14]]
        a=a+1
    wb.save(path+'\\protect\\static\\shuju\hrb.xlsx')
def zhengli():
    path = os.getcwd()
    wb = openpyxl.load_workbook(path+'\\protect\\static\\shuju\hrb.xlsx')
    ws = wb['Sheet1']
    a=ws.max_row
    for i in range(2,a+1):
        ws.cell(i,13).value=str(ws.cell(i,13).value).replace('元/平','').replace(',','')
        try:
            ws.cell(i, 6).value = str(ws.cell(i, 6).value).split(' ')[1]
        except:
            ws.cell(i, 6).value=''
    wb.save(path+'\\protect\\static\\shuju\hrb.xlsx')
def area():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\shuju\hrb.xlsx')
    pany_rank = pd.pivot_table(df, index='标签', values='名称', aggfunc='count').reset_index()
    value3 = pany_rank['名称']
    quxian  = pany_rank['标签']
    map3 = Map("济南地图", "济南", width='100%', height='900px')
    map3.add("济南", quxian, value3, visual_range=[min(value3), max(value3)], maptype='济南', is_visualmap=True)
    map3.render(path=path+"\\protect\\templates\\济南地图.html")
def danjia():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\shuju\hrb.xlsx')
    pany_rank = pd.pivot_table(df, index='标签', values='单价', aggfunc='mean').reset_index()
    value3 = pany_rank['单价']
    quxian  = pany_rank['标签']
    bar = Bar("二手房各区单价排名", width='100%', height='900px')
    bar.add("单价排名", quxian,  value3)
    bar.render(path=path+"\\protect\\templates\\二手房各区单价排名.html")
def pie():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\shuju\hrb.xlsx')
    pany_rank = pd.pivot_table(df, index='标签', values='名称', aggfunc='count').reset_index()
    value3 = pany_rank['名称'].to_list()
    quxian  = pany_rank['标签'].to_list()

    total = 0
    for ele in range(0, len(value3)):
        total = total + value3[ele]
    total_list=[]
    for ele in range(0, len(value3)):
        total_list.append(round(ele/total,2))
    pie = Pie("饼图", width='100%', height='900px')
    pie.add(
        "数量占比",
        quxian,
        value3,
        is_label_show=True,

    )
    pie.render(path=path+"\\protect\\templates\\各区二手房数量占比图.html")
def ciyun():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\shuju\hrb.xlsx')
    pany_rank = pd.pivot_table(df, index='小区名', values='名称', aggfunc='count').reset_index()
    value3 = pany_rank['名称']
    quxian  = pany_rank['小区名']
    wordcloud = WordCloud(width='100%', height='900px')
    wordcloud.add("小区词云图", quxian,value3, word_size_range=[min(value3),max(value3)])
    wordcloud.render(path=path + "\\protect\\templates\\小区词云图.html")

def nian():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\年份均价图.xlsx')

    quxian=df['年份'].to_list()
    value = df['均价'].to_list()
    bar = Bar("房间类型折线图",width='100%', height='900px')
    bar.add("单价", quxian, value)

    bar.render(path=path + "\\protect\\templates\\年份折线图1.html")
def chaoxiang():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\shuju\hrb.xlsx')
    # pany_rank = pd.pivot_table(df, index='朝向', values='单价', aggfunc='mean').reset_index()
    value3 = df['总价']
    quxian  = df['面积']
    scatter = Scatter("散点图", width='100%', height='900px')
    scatter.add("面积", quxian,value3
                )


    scatter.render(path=path+"\\protect\\templates\\二手房面积价格散点图.html")
def guanzhu():
    path = os.getcwd()
    df = pd.read_excel(path+'\\protect\\static\\shuju\hrb.xlsx')
    pany_rank = pd.pivot_table(df, index='标签', values='关注人数', aggfunc='sum').reset_index()
    value3 = pany_rank['关注人数']
    quxian  = pany_rank['标签']
    bar = Bar("二手房各区关注人数排名", width='100%', height='900px')
    bar.add("关注人数排名", quxian,  value3)
    bar.render(path=path+"\\protect\\templates\\二手房各区关注人数排名.html")
def get_price_forecast():
    path = os.getcwd()
    df = pd.read_excel(path + '\\protect\\static\\shuju\hrb.xlsx')
    pany_rank = pd.pivot_table(df, index='朝向', values='单价', aggfunc='count').reset_index()
    value3 = pany_rank['单价']
    quxian = pany_rank['朝向']
    bar = Bar("朝向饼图", width='100%', height='900px')
    bar.add(
        "朝向均价折线图",
        quxian,
        value3,
        is_label_show=True,
    )
    bar.render(path=path + "\\protect\\templates\\二手房朝向占比图.html")
    # data_copy = data.copy()
    # print(data_copy[['房间类型','面积']].head())
    # data_copy[['室','厅']]=data_copy['房间类型'].str.extract('(\d+)室(\d+)厅')
    # data_copy['室']=data_copy['室'].astype(float)
    # data_copy['厅']=data_copy['厅'].astype(float)#将房子厅转换为浮点类型data_copy['卫']= data_copy['卫'].astype(float)
    # print(data_copy[['室','厅']].head())
    # del data_copy['名称']
    # del data_copy['小区名']
    # del data_copy['位置']
    # del data_copy['房间类型']
    # del data_copy['朝向']
    # del data_copy['装修']
    # del data_copy['楼层']
    # del data_copy['年份']
    # del data_copy['楼况']
    # del data_copy['关注人数']
    # del data_copy['发布时间']
    # del data_copy['标签']
    # del data_copy['标签']
    # del data_copy['单价']
    #
    # data_copy.dropna(axis=0,how = 'any',inplace = True)  # 删除data 据中的所有空值
    # new_data = data_copy[data_copy['面积'] <300].reset_index(drop=True)
    # print(new_data.head())
    # # 添加自定义水刘数据
    # new_data.loc[2505] = [88.0,None,  2.0, 1.0]
    # new_data.loc[2506] = [ 136.0,None, 3.0, 2.0]
    # data_train = new_data.loc[0:2504]
    # x_list = ['面积','室','厅']
    # data_mean = data_train.mean()
    # data_std = data_train.std()
    # data_train = (data_train - data_mean) / data_std  # 数据标准化
    # x_train = data_train[x_list].values #特征数据
    # y_train = data_train['总价'].values #目标数据，总价
    # linearsvr =LinearSVR(C=0.1)#创建LinearSVR()对象
    # linearsvr.fit(x_train, y_train)
    # x = ((new_data[x_list] - data_mean[x_list]) / data_std[x_list]).values
    # new_data[u'y_pred'] = linearsvr.predict(x) * data_std['总价'] + data_mean['总价']
    # # new_data=new_data.sort_values(by="y_pred",ascending=False)
    # name=new_data['总价'].to_list()[0:10]
    # value = new_data['y_pred'].to_list()[0:10]
    # bar = Bar("二手房预测价格", width='100%', height='900px')
    # bar.add("二手房预测价格", value, name)
    # bar.render(path=path + "\\protect\\templates\\二手房预测价格.html")
    # print('真实值与预测值分别为：\n',new_data[['总价','y_pred']])
    # y = new_data[['总价']][2490:]
    # y_pred = new_data[['y_pred']][2490:]
    # print(y,y_pred)
    # return y, y_pred
ex()
zhengli()
area()
danjia()
pie()
ciyun()
nian()
# huxing()
chaoxiang()
guanzhu()
get_price_forecast()




